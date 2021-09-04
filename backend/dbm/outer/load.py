# Fernando Lavarreda
# Read beams and materials from files

import openpyxl as xl


def beam_xlsx(loctation:str, sheet:str, rows:list, columns:dict):
    assert len(columns) == 6
    wb = xl.load_workbook(loctation, True)
    beams = []
    if sheet in wb.sheetnames:
        file_sheet = wb[sheet]
        for row in rows:
            name = file_sheet.cell(row, columns["name"]).value
            type_ = file_sheet.cell(row, columns["type"]).value
            inertia = int(file_sheet.cell(row, columns["inertia"]).value*1000000)
            area = int(file_sheet.cell(row, columns["area"]).value)
            height = int(file_sheet.cell(row, columns["height"]).value)
            weight = int(file_sheet.cell(row, columns["weight"]).value)
            beams.append((name, type_, inertia, area, height, weight))
    return beams


def material_xlsx(loctation:str, sheet:str, rows:list, columns:dict):
    assert len(columns) == 4
    wb = xl.load_workbook(loctation, True)
    materials = []
    if sheet in wb.sheetnames:
        file_sheet = wb[sheet]
        for row in rows:
            name = file_sheet.cell(row, columns["name"]).value
            ymodulus = float(file_sheet.cell(row, columns["ymodulus"]).value)
            comp_stress = float(file_sheet.cell(row, columns["comp_stress"]).value)
            ten_stress = float(file_sheet.cell(row, columns["ten_stress"]).value)
            materials.append((name, ymodulus, comp_stress, ten_stress))
    return materials